#!/usr/bin/env python3
"""
CH Construct Timesheet - Excel Offerte Generator
Genereert professioneel Excel pricing sheet met 8 worksheets
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from datetime import datetime

# Kleuren schema (Bouw/Tech - Blauw/Oranje/Grijs)
COLORS = {
    'primary': '1E3A8A',      # Donkerblauw
    'secondary': 'F97316',    # Oranje
    'accent': '64748B',       # Grijs
    'success': '10B981',      # Groen
    'warning': 'F59E0B',      # Geel
    'danger': 'EF4444',       # Rood
    'light': 'F1F5F9',        # Licht grijs
    'white': 'FFFFFF',
    'header': '0F172A',       # Zeer donker blauw
}

def create_header_style():
    """Stijl voor hoofdheaders"""
    return {
        'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            bottom=Side(style='medium', color='000000')
        )
    }

def create_subheader_style():
    """Stijl voor subheaders"""
    return {
        'font': Font(name='Calibri', size=12, bold=True, color='000000'),
        'fill': PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center'),
    }

def create_currency_format():
    """Excel currency format"""
    return '€#,##0.00'

def apply_cell_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Helper om stijlen toe te passen"""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format

def create_workbook():
    """Creëer het Excel workbook met alle sheets"""
    wb = Workbook()
    
    # Verwijder default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Creëer alle sheets
    create_executive_summary(wb)
    create_development_costs(wb)
    create_pricing_calculator(wb)
    create_roi_analysis(wb)
    create_competitive_comparison(wb)
    create_roadmap_timeline(wb)
    create_maintenance_support(wb)
    create_license_comparison(wb)
    
    return wb

def create_executive_summary(wb):
    """Sheet 1: Executive Summary"""
    ws = wb.create_sheet("Executive Summary", 0)
    
    # Header
    ws['A1'] = 'CH CONSTRUCT TIMESHEET'
    ws['A1'].font = Font(name='Calibri', size=20, bold=True, color=COLORS['primary'])
    ws.merge_cells('A1:F1')
    
    ws['A2'] = 'COMMERCIËLE OFFERTE'
    ws['A2'].font = Font(name='Calibri', size=16, bold=True, color=COLORS['secondary'])
    ws.merge_cells('A2:F2')
    
    # Metadata
    row = 4
    metadata = [
        ('Versie:', '1.0'),
        ('Datum:', datetime.now().strftime('%d %B %Y')),
        ('Sector:', 'Bouw & Constructie | Technische Dienstverlening'),
        ('Type:', 'Progressive Web Application (PWA)'),
    ]
    
    for label, value in metadata:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Value Propositions
    row += 2
    ws[f'A{row}'] = 'UNIQUE VALUE PROPOSITIONS'
    header_style = create_header_style()
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    vps = [
        '✅ 100% Offline Functionaliteit - Werkt overal, ook op de bouwplaats',
        '✅ Privacy First - Alle data blijft lokaal op het apparaat',
        '✅ Bouw-specifiek - Chauffeur status, pauze, locatie per klant',
        '✅ Multi-profiel Support - Meerdere werknemers op één apparaat',
        '✅ Professionele PDF Export - Direct verstuurklare urenstaten',
        '✅ Nederlandse Interface - Volledig Nederlands, intuïtief',
        '✅ Geen SaaS verplicht - Keuze tussen modellen',
    ]
    
    for vp in vps:
        ws[f'A{row}'] = vp
        ws[f'A{row}'].font = Font(name='Calibri', size=10)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
    
    # Kosten samenvatting
    row += 2
    ws[f'A{row}'] = 'KOSTENSAMENVATTING'
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    headers = ['Component', 'Uren', 'Tarief', 'Ex BTW', 'Inc BTW (21%)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        subheader_style = create_subheader_style()
        apply_cell_style(cell, **subheader_style)
    
    row += 1
    cost_data = [
        ('Basis App Development', 85, 75, '=B{0}*C{0}', '=D{0}*1.21'),
        ('Project Management', 10, 75, '=B{0}*C{0}', '=D{0}*1.21'),
        ('QA & Testing', 8, 65, '=B{0}*C{0}', '=D{0}*1.21'),
        ('Performance Optimalisatie', 6, 85, '=B{0}*C{0}', '=D{0}*1.21'),
        ('Documentatie', 4, 65, '=B{0}*C{0}', '=D{0}*1.21'),
    ]
    
    start_row = row
    for item, uren, tarief, ex_formula, inc_formula in cost_data:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = uren
        ws[f'C{row}'] = f'€{tarief}'
        ws[f'D{row}'] = ex_formula.format(row)
        ws[f'E{row}'] = inc_formula.format(row)
        
        ws[f'D{row}'].number_format = create_currency_format()
        ws[f'E{row}'].number_format = create_currency_format()
        row += 1
    
    # Totaal
    ws[f'A{row}'] = 'TOTAAL BASIS APP'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'D{row}'] = f'=SUM(D{start_row}:D{row-1})'
    ws[f'E{row}'] = f'=SUM(E{start_row}:E{row-1})'
    ws[f'D{row}'].number_format = create_currency_format()
    ws[f'E{row}'].number_format = create_currency_format()
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'E{row}'].font = Font(bold=True)
    
    # Uitbreidingen samenvatting
    row += 3
    ws[f'A{row}'] = 'POTENTIËLE UITBREIDINGEN (Optioneel)'
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    uitbreidingen = [
        ('Backend/API Integratie', 100, 7500),
        ('Multi-user/Team Features', 80, 6000),
        ('iOS Native App', 140, 10500),
        ('Android Native App', 120, 9000),
        ('HR/Salaris Integraties', 60, 4500),
        ('Advanced Reporting', 80, 6000),
        ('White-label Branding', 50, 3750),
    ]
    
    for module, uren, kosten in uitbreidingen:
        ws[f'A{row}'] = module
        ws[f'B{row}'] = uren
        ws[f'D{row}'] = kosten
        ws[f'E{row}'] = f'=D{row}*1.21'
        ws[f'D{row}'].number_format = create_currency_format()
        ws[f'E{row}'].number_format = create_currency_format()
        row += 1
    
    # Aanbeveling
    row += 2
    ws[f'A{row}'] = 'AANBEVOLEN PAKKET'
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Voor Bouw & Technische Dienstverlening (10-50 werknemers):'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Hybrid Model: €4,500 setup + €6/gebruiker/maand'
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws[f'A{row}'] = 'Inclusief: Training, support, updates, backups'
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws[f'A{row}'] = '3-jaar TCO (20 gebruikers): €13,860'
    ws[f'A{row}'].font = Font(bold=True, color=COLORS['success'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws[f'A{row}'] = 'ROI: 10-12 maanden'
    ws[f'A{row}'].font = Font(bold=True, color=COLORS['success'])
    ws.merge_cells(f'A{row}:F{row}')
    
    # Kolom breedtes
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18

def create_development_costs(wb):
    """Sheet 2: Ontwikkelkosten Breakdown"""
    ws = wb.create_sheet("Ontwikkelkosten")
    
    # Header
    ws['A1'] = 'ONTWIKKELKOSTEN BREAKDOWN'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=COLORS['primary'])
    ws.merge_cells('A1:G1')
    
    # Headers
    row = 3
    headers = ['Module/Feature', 'Complexiteit', 'Uren (Low)', 'Uren (High)', 'Tarief €/u', 'Kosten (ex BTW)', 'Kosten (inc BTW)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        header_style = create_header_style()
        apply_cell_style(cell, **header_style)
    
    # FASE 1: Huidige App
    row += 1
    ws[f'A{row}'] = 'FASE 1: CORE APP (VOLTOOID ✅)'
    ws[f'A{row}'].font = Font(bold=True, size=12, color=COLORS['success'])
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    fase1_data = [
        ('Database Setup (Dexie/IndexedDB)', 'Medium', 4, 6, 75),
        ('Time Entry CRUD Operations', 'Medium', 6, 10, 75),
        ('Multi-profiel Systeem', 'Medium', 5, 8, 75),
        ('Client & Location Management', 'Laag', 3, 5, 65),
        ('Week Overzicht & Navigatie', 'Medium', 5, 8, 75),
        ('PDF Export Functionaliteit', 'Hoog', 8, 12, 85),
        ('Logo Management & Resizing', 'Medium', 3, 5, 75),
        ('Data Import/Export/Backup', 'Medium', 5, 8, 75),
        ('Notification Systeem', 'Medium', 4, 6, 75),
        ('Service Worker (PWA)', 'Medium', 3, 5, 75),
        ('Swipe Gestures', 'Laag', 2, 3, 65),
        ('Migration Systeem', 'Laag', 2, 4, 65),
        ('UI/UX Design & Styling', 'Medium', 6, 10, 75),
        ('Testing & Debugging', 'Standaard', 4, 8, 65),
        ('Performance Optimalisatie', 'Medium', 4, 6, 85),
        ('Code Review & Refactoring', 'Standaard', 2, 4, 75),
        ('Documentatie', 'Standaard', 3, 5, 65),
        ('Project Management', 'Standaard', 8, 12, 75),
    ]
    
    start_row = row
    for feature, complexity, low, high, tarief in fase1_data:
        ws[f'A{row}'] = feature
        ws[f'B{row}'] = complexity
        ws[f'C{row}'] = low
        ws[f'D{row}'] = high
        ws[f'E{row}'] = tarief
        ws[f'F{row}'] = f'=ROUND(((C{row}+D{row})/2)*E{row}, 0)'
        ws[f'G{row}'] = f'=F{row}*1.21'
        
        ws[f'E{row}'].number_format = '€#,##0'
        ws[f'F{row}'].number_format = create_currency_format()
        ws[f'G{row}'].number_format = create_currency_format()
        
        # Color code complexity
        if complexity == 'Hoog':
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['danger'], end_color=COLORS['danger'], fill_type='solid')
            ws[f'B{row}'].font = Font(color='FFFFFF', bold=True)
        elif complexity == 'Medium':
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
        
        row += 1
    
    # Subtotaal FASE 1
    ws[f'A{row}'] = 'SUBTOTAAL FASE 1'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'C{row}'] = f'=SUM(C{start_row}:C{row-1})'
    ws[f'D{row}'] = f'=SUM(D{start_row}:D{row-1})'
    ws[f'F{row}'] = f'=SUM(F{start_row}:F{row-1})'
    ws[f'G{row}'] = f'=SUM(G{start_row}:G{row-1})'
    
    for col in ['C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True)
    for col in ['F', 'G']:
        ws[f'{col}{row}'].number_format = create_currency_format()
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
        ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
    
    # FASE 2: Uitbreidingen
    row += 3
    ws[f'A{row}'] = 'FASE 2: BACKEND & MULTI-USER (Optioneel)'
    ws[f'A{row}'].font = Font(bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    fase2_data = [
        ('REST API Setup (Node.js/Firebase)', 'Hoog', 20, 30, 85),
        ('Authentication & Authorization', 'Hoog', 15, 20, 85),
        ('Data Sync Protocollen', 'Hoog', 15, 25, 85),
        ('Real-time Updates', 'Hoog', 10, 15, 85),
        ('Team Dashboard', 'Medium', 10, 15, 75),
        ('Manager Goedkeuringen', 'Medium', 8, 12, 75),
        ('Role-based Access Control', 'Hoog', 8, 12, 85),
        ('Activity Logging', 'Laag', 4, 6, 65),
        ('Integration Testing', 'Medium', 8, 12, 75),
        ('Load Testing (100+ users)', 'Medium', 5, 8, 75),
        ('Security Audit', 'Hoog', 8, 12, 100),
        ('Production Deployment', 'Medium', 4, 6, 75),
    ]
    
    start_row2 = row
    for feature, complexity, low, high, tarief in fase2_data:
        ws[f'A{row}'] = feature
        ws[f'B{row}'] = complexity
        ws[f'C{row}'] = low
        ws[f'D{row}'] = high
        ws[f'E{row}'] = tarief
        ws[f'F{row}'] = f'=ROUND(((C{row}+D{row})/2)*E{row}, 0)'
        ws[f'G{row}'] = f'=F{row}*1.21'
        
        ws[f'E{row}'].number_format = '€#,##0'
        ws[f'F{row}'].number_format = create_currency_format()
        ws[f'G{row}'].number_format = create_currency_format()
        
        if complexity == 'Hoog':
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['danger'], end_color=COLORS['danger'], fill_type='solid')
            ws[f'B{row}'].font = Font(color='FFFFFF', bold=True)
        elif complexity == 'Medium':
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
        
        row += 1
    
    # Subtotaal FASE 2
    ws[f'A{row}'] = 'SUBTOTAAL FASE 2'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'C{row}'] = f'=SUM(C{start_row2}:C{row-1})'
    ws[f'D{row}'] = f'=SUM(D{start_row2}:D{row-1})'
    ws[f'F{row}'] = f'=SUM(F{start_row2}:F{row-1})'
    ws[f'G{row}'] = f'=SUM(G{start_row2}:G{row-1})'
    
    for col in ['C', 'D']:
        ws[f'{col}{row}'].font = Font(bold=True)
    for col in ['F', 'G']:
        ws[f'{col}{row}'].number_format = create_currency_format()
        ws[f'{col}{row}'].font = Font(bold=True)
    
    # Kolom breedtes
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20

def create_pricing_calculator(wb):
    """Sheet 3: Pricing Calculator (Interactief)"""
    ws = wb.create_sheet("Pricing Calculator")
    
    # Header
    ws['A1'] = 'PRICING CALCULATOR'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=COLORS['primary'])
    ws.merge_cells('A1:E1')
    
    # Input sectie
    row = 3
    ws[f'A{row}'] = 'INPUT PARAMETERS'
    header_style = create_header_style()
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 2
    ws[f'A{row}'] = 'Aantal Gebruikers:'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = 10
    ws[f'B{row}'].number_format = '0'
    ws[f'B{row}'].fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
    
    row += 1
    ws[f'A{row}'] = 'Contract Periode (jaren):'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = 3
    ws[f'B{row}'].number_format = '0'
    ws[f'B{row}'].fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
    
    row += 1
    ws[f'A{row}'] = 'Support Level:'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = 'Professional'
    ws[f'B{row}'].fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
    
    # Bewaar input cell references
    users_cell = f'B{row-2}'
    years_cell = f'B{row-1}'
    support_cell = f'B{row}'
    
    # Pricing Models Comparison
    row += 3
    ws[f'A{row}'] = 'PRICING MODELS VERGELIJKING'
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = 'Model'
    ws[f'B{row}'] = 'Jaar 1'
    ws[f'C{row}'] = 'Jaar 2'
    ws[f'D{row}'] = 'Jaar 3'
    ws[f'E{row}'] = '3-jaar Totaal'
    
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        subheader_style = create_subheader_style()
        apply_cell_style(cell, **subheader_style)
    
    row += 1
    
    # Model A: SaaS
    ws[f'A{row}'] = 'A. SaaS Subscription (€15/user/maand)'
    monthly_rate = 15
    ws[f'B{row}'] = f'={users_cell}*{monthly_rate}*12'
    ws[f'C{row}'] = f'=B{row}'
    ws[f'D{row}'] = f'=B{row}'
    ws[f'E{row}'] = f'=B{row}+C{row}+D{row}'
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = create_currency_format()
    
    row += 1
    
    # Model B: One-time
    ws[f'A{row}'] = 'B. One-time License + Onderhoud'
    ws[f'B{row}'] = 5000  # Base license
    ws[f'C{row}'] = f'=B{row}*0.20'  # 20% onderhoud
    ws[f'D{row}'] = f'=B{row}*0.20'
    ws[f'E{row}'] = f'=B{row}+C{row}+D{row}'
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = create_currency_format()
    
    row += 1
    
    # Model C: Hybrid (AANBEVOLEN)
    ws[f'A{row}'] = 'C. Hybrid (€4,500 + €6/user/maand) ⭐'
    ws[f'A{row}'].font = Font(bold=True, color=COLORS['success'])
    ws[f'B{row}'] = f'=4500+{users_cell}*6*12'
    ws[f'C{row}'] = f'={users_cell}*6*12'
    ws[f'D{row}'] = f'={users_cell}*6*12'
    ws[f'E{row}'] = f'=B{row}+C{row}+D{row}'
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = create_currency_format()
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
        ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
    
    # Volume Discount Matrix
    row += 4
    ws[f'A{row}'] = 'VOLUME DISCOUNT TABEL'
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = 'Gebruikers'
    ws[f'B{row}'] = 'Discount %'
    ws[f'C{row}'] = 'SaaS €/maand'
    ws[f'D{row}'] = 'Hybrid Setup'
    
    row += 1
    discount_data = [
        ('1-5', 0, 15, 2500),
        ('6-15', 5, 14.25, 4000),
        ('16-25', 10, 13.50, 4500),
        ('26-50', 15, 12.75, 5500),
        ('50+', 20, 12.00, 7000),
    ]
    
    for users_range, discount, saas, hybrid in discount_data:
        ws[f'A{row}'] = users_range
        ws[f'B{row}'] = f'{discount}%'
        ws[f'C{row}'] = saas
        ws[f'D{row}'] = hybrid
        ws[f'C{row}'].number_format = '€#,##0.00'
        ws[f'D{row}'].number_format = '€#,##0'
        row += 1
    
    # Kolom breedtes
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18

def create_roi_analysis(wb):
    """Sheet 4: ROI Analysis"""
    ws = wb.create_sheet("ROI Analysis")
    
    # Header
    ws['A1'] = 'ROI ANALYSIS VOOR KLANTEN'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=COLORS['primary'])
    ws.merge_cells('A1:F1')
    
    # Scenario parameters
    row = 3
    ws[f'A{row}'] = 'SCENARIO: BOUWBEDRIJF MET 10 WERKNEMERS'
    header_style = create_header_style()
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    ws[f'A{row}'] = 'Parameter'
    ws[f'B{row}'] = 'Waarde'
    ws[f'C{row}'] = 'Eenheid'
    
    row += 1
    params = [
        ('Aantal werknemers', 10, 'personen'),
        ('Gemiddeld uurloon', 30, '€/uur'),
        ('Werkweken per jaar', 48, 'weken'),
    ]
    
    for param, value, unit in params:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = value
        ws[f'C{row}'] = unit
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Huidige situatie
    row += 2
    ws[f'A{row}'] = 'HUIDIGE SITUATIE (Papieren urenstaten)'
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Tijd per werknemer per week'
    ws[f'B{row}'] = 0.5
    ws[f'C{row}'] = 'uur'
    ws[f'A{row}'].font = Font(bold=True)
    
    current_time_cell = f'B{row}'
    
    row += 1
    ws[f'A{row}'] = 'Totale tijd per week'
    ws[f'B{row}'] = f'=B{row-4}*{current_time_cell}'  # 10 * 0.5
    ws[f'C{row}'] = 'uur'
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = 'Maandelijkse kosten (4 weken)'
    ws[f'B{row}'] = f'=B{row-1}*4*B{row-5}'  # totale uren * 4 * €30
    ws[f'C{row}'] = '€/maand'
    ws[f'B{row}'].number_format = create_currency_format()
    ws[f'A{row}'].font = Font(bold=True)
    
    current_monthly_cell = f'B{row}'
    
    row += 1
    ws[f'A{row}'] = 'Jaarlijkse kosten'
    ws[f'B{row}'] = f'={current_monthly_cell}*12'
    ws[f'C{row}'] = '€/jaar'
    ws[f'B{row}'].number_format = create_currency_format()
    ws[f'B{row}'].fill = PatternFill(start_color=COLORS['danger'], end_color=COLORS['danger'], fill_type='solid')
    ws[f'B{row}'].font = Font(bold=True, color='FFFFFF')
    
    current_yearly_cell = f'B{row}'
    
    # Met CH Construct Timesheet
    row += 3
    ws[f'A{row}'] = 'MET CH CONSTRUCT TIMESHEET'
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Tijd per werknemer per week'
    ws[f'B{row}'] = 0.083  # 5 minuten
    ws[f'C{row}'] = 'uur (5 min)'
    ws[f'A{row}'].font = Font(bold=True)
    
    new_time_cell = f'B{row}'
    
    row += 1
    ws[f'A{row}'] = 'Totale tijd per week'
    ws[f'B{row}'] = f'=B{row-14}*{new_time_cell}'  # 10 * 0.083
    ws[f'C{row}'] = 'uur'
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = 'Maandelijkse besparing'
    ws[f'B{row}'] = f'={current_monthly_cell}-B{row-1}*4*B{row-15}'
    ws[f'C{row}'] = '€/maand'
    ws[f'B{row}'].number_format = create_currency_format()
    ws[f'A{row}'].font = Font(bold=True)
    
    monthly_savings_cell = f'B{row}'
    
    row += 1
    ws[f'A{row}'] = 'Jaarlijkse besparing'
    ws[f'B{row}'] = f'={monthly_savings_cell}*12'
    ws[f'C{row}'] = '€/jaar'
    ws[f'B{row}'].number_format = create_currency_format()
    ws[f'B{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    ws[f'B{row}'].font = Font(bold=True, color='FFFFFF')
    
    yearly_savings_cell = f'B{row}'
    
    # Investment & ROI
    row += 3
    ws[f'A{row}'] = 'INVESTERING & ROI (Hybrid Model)'
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Setup kosten'
    ws[f'B{row}'] = 4500
    ws[f'C{row}'] = '€ (eenmalig)'
    ws[f'B{row}'].number_format = create_currency_format()
    
    row += 1
    ws[f'A{row}'] = 'Jaarlijkse kosten'
    ws[f'B{row}'] = '=10*6*12'  # 10 users * €6/maand
    ws[f'C{row}'] = '€/jaar'
    ws[f'B{row}'].number_format = create_currency_format()
    
    yearly_cost_cell = f'B{row}'
    
    row += 1
    ws[f'A{row}'] = '3-jaar totale investering'
    ws[f'B{row}'] = f'=B{row-2}+{yearly_cost_cell}*3'
    ws[f'C{row}'] = '€ (3 jaar)'
    ws[f'B{row}'].number_format = create_currency_format()
    
    total_investment_cell = f'B{row}'
    
    row += 2
    ws[f'A{row}'] = '3-jaar besparing'
    ws[f'B{row}'] = f'={yearly_savings_cell}*3'
    ws[f'C{row}'] = '€'
    ws[f'B{row}'].number_format = create_currency_format()
    ws[f'B{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    
    total_savings_cell = f'B{row}'
    
    row += 1
    ws[f'A{row}'] = 'Netto winst (3 jaar)'
    ws[f'B{row}'] = f'={total_savings_cell}-{total_investment_cell}'
    ws[f'C{row}'] = '€'
    ws[f'B{row}'].number_format = create_currency_format()
    ws[f'B{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    ws[f'B{row}'].font = Font(bold=True, color='FFFFFF', size=12)
    
    row += 1
    ws[f'A{row}'] = 'ROI Percentage'
    ws[f'B{row}'] = f'=({total_savings_cell}-{total_investment_cell})/{total_investment_cell}'
    ws[f'C{row}'] = '%'
    ws[f'B{row}'].number_format = '0%'
    ws[f'B{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    ws[f'B{row}'].font = Font(bold=True, color='FFFFFF', size=12)
    
    row += 1
    ws[f'A{row}'] = 'Break-even periode'
    ws[f'B{row}'] = f'=B{row-7}/({yearly_savings_cell}/12)'
    ws[f'C{row}'] = 'maanden'
    ws[f'B{row}'].number_format = '0.0'
    ws[f'B{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    ws[f'B{row}'].font = Font(bold=True, color='FFFFFF', size=12)
    
    # Extra voordelen
    row += 3
    ws[f'A{row}'] = 'EXTRA VOORDELEN (Niet gekwantificeerd)'
    apply_cell_style(ws[f'A{row}'], **header_style)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    benefits = [
        '✅ Minder fouten in urenstaten (-25%)',
        '✅ Snellere facturatie (3-5 dagen eerder)',
        '✅ Betere data voor projectmanagement',
        '✅ Professionelere uitstraling naar klanten',
        '✅ Compliance & audit trail',
        '✅ Minder telefoontjes "waar is je urenstaat?"',
    ]
    
    for benefit in benefits:
        ws[f'A{row}'] = benefit
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
    
    # Kolom breedtes
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

def create_competitive_comparison(wb):
    """Sheet 5: Competitive Comparison"""
    ws = wb.create_sheet("Concurrentie")
    
    # Header
    ws['A1'] = 'CONCURRENTIE ANALYSE'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=COLORS['primary'])
    ws.merge_cells('A1:H1')
    
    # Headers
    row = 3
    headers = ['Feature/Product', 'Toggl Track', 'Clockify', 'Harvest', 'TimeCamp', 'CH Construct', 'Gewicht']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        header_style = create_header_style()
        apply_cell_style(cell, **header_style)
    
    # Features comparison
    row += 1
    features = [
        ('Prijs/gebruiker/maand', '€9-18', '€0-10', '€12-16', '€7-10', '€6-25*', 20),
        ('Offline functionaliteit', '❌ Beperkt', '❌ Nee', '❌ Nee', '⚠️ Beperkt', '✅ Volledig', 25),
        ('Data privacy (lokaal)', '❌ Cloud', '❌ Cloud', '❌ Cloud', '❌ Cloud', '✅ 100% Lokaal', 15),
        ('Bouw-specifieke features', '❌ Nee', '❌ Nee', '❌ Nee', '❌ Nee', '✅ Ja', 20),
        ('Nederlandse interface', '⚠️ Engels', '⚠️ Engels', '❌ Engels', '⚠️ Engels', '✅ Nederlands', 10),
        ('Multi-profiel (1 device)', '❌ Nee', '❌ Nee', '❌ Nee', '❌ Nee', '✅ Ja', 15),
        ('PDF met logo', '⚠️ Basis', '⚠️ Basis', '✅ Goed', '⚠️ Basis', '✅ Professioneel', 10),
        ('Setup tijd', '2-3 uur', '2-3 uur', '3-4 uur', '2-3 uur', '10 min', 5),
        ('Zonder internet vereist', '❌ Nee', '❌ Nee', '❌ Nee', '❌ Nee', '✅ Ja', 25),
        ('Pauze registratie', '✅ Ja', '✅ Ja', '✅ Ja', '✅ Ja', '✅ Ja + Auto', 5),
    ]
    
    for feature, toggl, clockify, harvest, timecamp, ch_construct, weight in features:
        ws[f'A{row}'] = feature
        ws[f'B{row}'] = toggl
        ws[f'C{row}'] = clockify
        ws[f'D{row}'] = harvest
        ws[f'E{row}'] = timecamp
        ws[f'F{row}'] = ch_construct
        ws[f'G{row}'] = weight
        
        # Highlight CH Construct column
        ws[f'F{row}'].fill = PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid')
        
        # Color code based on symbols
        for col in ['B', 'C', 'D', 'E', 'F']:
            cell_value = ws[f'{col}{row}'].value
            if isinstance(cell_value, str):
                if '✅' in cell_value:
                    ws[f'{col}{row}'].font = Font(color=COLORS['success'], bold=True)
                elif '❌' in cell_value:
                    ws[f'{col}{row}'].font = Font(color=COLORS['danger'])
                elif '⚠️' in cell_value:
                    ws[f'{col}{row}'].font = Font(color=COLORS['warning'])
        
        row += 1
    
    # Score berekening
    row += 2
    ws[f'A{row}'] = 'TOTAAL SCORE (gewogen)'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    
    # Manual scores (simplified for demo)
    scores = [
        ('Toggl', 65),
        ('Clockify', 70),
        ('Harvest', 68),
        ('TimeCamp', 72),
        ('CH Construct', 92),
    ]
    
    row += 1
    for product, score in scores:
        ws[f'A{row}'] = product
        ws[f'B{row}'] = score
        ws[f'C{row}'] = '/100'
        
        if product == 'CH Construct':
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
            ws[f'B{row}'].font = Font(bold=True, color='FFFFFF', size=12)
        
        row += 1
    
    # Unique Selling Points
    row += 2
    ws[f'A{row}'] = 'UNIQUE SELLING POINTS CH CONSTRUCT'
    apply_cell_style(ws[f'A{row}'], **create_header_style())
    ws.merge_cells(f'A{row}:H{row}')
    
    row += 1
    usps = [
        '🎯 Enige oplossing specifiek voor Nederlandse bouw/techniek sector',
        '📱 Werkt 100% offline - geen internet vereist op de bouwplaats',
        '🔒 Privacy first - alle data blijft lokaal (GDPR compliant)',
        '👥 Multi-profiel support - meerdere werknemers op één tablet',
        '🇳🇱 Volledig Nederlandse interface - intuïtief voor werkvloer',
        '💰 Geen SaaS lock-in - keuze tussen pricing modellen',
        '⚡ Ultrasnelle setup - 10 minuten vs. 2-3 uur bij concurrenten',
    ]
    
    for usp in usps:
        ws[f'A{row}'] = usp
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].font = Font(size=10)
        row += 1
    
    # Kolom breedtes
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['G'].width = 12

def create_roadmap_timeline(wb):
    """Sheet 6: Roadmap & Timeline"""
    ws = wb.create_sheet("Roadmap")
    
    # Header
    ws['A1'] = 'DEVELOPMENT ROADMAP & TIMELINE'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=COLORS['primary'])
    ws.merge_cells('A1:F1')
    
    # FASE 1
    row = 3
    ws[f'A{row}'] = 'FASE 1: CORE APP (VOLTOOID ✅)'
    ws[f'A{row}'].font = Font(bold=True, size=12, color=COLORS['success'])
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    ws[f'A{row}'].font = Font(bold=True, color='FFFFFF')
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Duur: 5 dagen (10-15 april 2026)'
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Status: Production Ready | Kosten: €8,415 (ex BTW)'
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    fase1_deliverables = [
        '✅ Database setup (Dexie/IndexedDB)',
        '✅ Time entry CRUD operaties',
        '✅ Multi-profiel management',
        '✅ Client & location management',
        '✅ Week overzicht & navigatie',
        '✅ PDF export met logo',
        '✅ PWA installatie & offline mode',
        '✅ Notification systeem',
        '✅ Performance optimalisaties',
        '✅ Volledige documentatie',
    ]
    
    for deliverable in fase1_deliverables:
        ws[f'A{row}'] = deliverable
        ws[f'A{row}'].font = Font(color=COLORS['success'])
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
    
    # FASE 2
    row += 2
    ws[f'A{row}'] = 'FASE 2: BACKEND & MULTI-USER'
    ws[f'A{row}'].font = Font(bold=True, size=12, color=COLORS['primary'])
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    ws[f'A{row}'].font = Font(bold=True, color='FFFFFF')
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Duur: 8-12 weken | Kosten: €7,500 (ex BTW)'
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    ws[f'A{row}'] = 'Week 1-3: API Development'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = '  • REST API setup (Node.js/Firebase)'
    row += 1
    ws[f'A{row}'] = '  • Authentication & authorization'
    row += 1
    ws[f'A{row}'] = '  • Data sync protocollen'
    
    row += 2
    ws[f'A{row}'] = 'Week 4-6: Multi-user Features'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = '  • Team dashboard'
    row += 1
    ws[f'A{row}'] = '  • Manager goedkeuringen'
    row += 1
    ws[f'A{row}'] = '  • Role-based access control'
    
    row += 2
    ws[f'A{row}'] = 'Week 7-8: Testing & Deployment'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = '  • Integration testing'
    row += 1
    ws[f'A{row}'] = '  • Load testing (100+ users)'
    row += 1
    ws[f'A{row}'] = '  • Security audit'
    
    # FASE 3
    row += 3
    ws[f'A{row}'] = 'FASE 3: NATIVE MOBILE APPS'
    ws[f'A{row}'].font = Font(bold=True, size=12, color=COLORS['primary'])
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    ws[f'A{row}'].font = Font(bold=True, color='FFFFFF')
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Duur: 12-16 weken | Kosten: €19,500 (iOS €10,500 + Android €9,000)'
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    ws[f'A{row}'] = 'Week 1-8: iOS Development'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = '  • App structure & navigation'
    row += 1
    ws[f'A{row}'] = '  • Core features port'
    row += 1
    ws[f'A{row}'] = '  • iOS-specific features (Face ID, widgets)'
    row += 1
    ws[f'A{row}'] = '  • App Store submission'
    
    row += 2
    ws[f'A{row}'] = 'Week 1-6: Android Development (parallel)'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = '  • App structure & navigation'
    row += 1
    ws[f'A{row}'] = '  • Core features port'
    row += 1
    ws[f'A{row}'] = '  • Play Store submission'
    
    # Summary table
    row += 4
    ws[f'A{row}'] = 'SAMENVATTING PLANNING'
    apply_cell_style(ws[f'A{row}'], **create_header_style())
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Fase'
    ws[f'B{row}'] = 'Duur'
    ws[f'C{row}'] = 'Status'
    ws[f'D{row}'] = 'Kosten (ex BTW)'
    ws[f'E{row}'] = 'Kosten (inc BTW)'
    
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        subheader_style = create_subheader_style()
        apply_cell_style(cell, **subheader_style)
    
    row += 1
    summary_data = [
        ('Fase 1: Core App', '5 dagen', '✅ Voltooid', 8415, '=D{0}*1.21'),
        ('Fase 2: Backend', '8-12 weken', '📅 Te plannen', 7500, '=D{0}*1.21'),
        ('Fase 3: Mobile Apps', '12-16 weken', '📅 Te plannen', 19500, '=D{0}*1.21'),
        ('Fase 4: Enterprise', '6-10 weken', '📅 Te plannen', 14250, '=D{0}*1.21'),
    ]
    
    start_row = row
    for fase, duur, status, kosten, inc_formula in summary_data:
        ws[f'A{row}'] = fase
        ws[f'B{row}'] = duur
        ws[f'C{row}'] = status
        ws[f'D{row}'] = kosten
        ws[f'E{row}'] = inc_formula.format(row)
        
        ws[f'D{row}'].number_format = create_currency_format()
        ws[f'E{row}'].number_format = create_currency_format()
        
        if '✅' in status:
            ws[f'C{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
            ws[f'C{row}'].font = Font(color='FFFFFF', bold=True)
        
        row += 1
    
    # Totaal
    ws[f'A{row}'] = 'TOTAAL ALLE FASES'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'D{row}'] = f'=SUM(D{start_row}:D{row-1})'
    ws[f'E{row}'] = f'=SUM(E{start_row}:E{row-1})'
    ws[f'D{row}'].number_format = create_currency_format()
    ws[f'E{row}'].number_format = create_currency_format()
    ws[f'D{row}'].fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    ws[f'E{row}'].fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    ws[f'D{row}'].font = Font(bold=True, color='FFFFFF')
    ws[f'E{row}'].font = Font(bold=True, color='FFFFFF')
    
    # Kolom breedtes
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20

def create_maintenance_support(wb):
    """Sheet 7: Maintenance & Support Costs"""
    ws = wb.create_sheet("Onderhoud & Support")
    
    # Header
    ws['A1'] = 'ONDERHOUD & SUPPORT KOSTEN'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=COLORS['primary'])
    ws.merge_cells('A1:E1')
    
    # Support tiers
    row = 3
    ws[f'A{row}'] = 'JAARLIJKSE SUPPORT PAKKETTEN'
    apply_cell_style(ws[f'A{row}'], **create_header_style())
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    headers = ['Tier', 'Prijs/jaar (ex BTW)', 'Response Time', 'Kanalen', 'SLA Uptime']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        subheader_style = create_subheader_style()
        apply_cell_style(cell, **subheader_style)
    
    row += 1
    tiers = [
        ('Basic', 1200, '48 uur', 'Email', '99%'),
        ('Professional', 2500, '24 uur', 'Email + Telefoon', '99.5%'),
        ('Enterprise', 5000, '4 uur (24/7)', 'All + Dedicated', '99.9%'),
    ]
    
    for tier, price, response, channels, sla in tiers:
        ws[f'A{row}'] = tier
        ws[f'B{row}'] = price
        ws[f'C{row}'] = response
        ws[f'D{row}'] = channels
        ws[f'E{row}'] = sla
        
        ws[f'B{row}'].number_format = create_currency_format()
        
        if tier == 'Professional':
            ws[f'A{row}'].fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
            ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
    
    # Inclusief in alle tiers
    row += 2
    ws[f'A{row}'] = 'INCLUSIEF IN ALLE TIERS'
    apply_cell_style(ws[f'A{row}'], **create_header_style())
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    inclusions = [
        '✅ Security updates & patches',
        '✅ Browser compatibility updates',
        '✅ Bug fixes (critical binnen response time)',
        '✅ Minor feature updates',
        '✅ Email support',
        '✅ Quarterly health check',
        '✅ Documentation updates',
    ]
    
    for inclusion in inclusions:
        ws[f'A{row}'] = inclusion
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
    
    # Professional & Enterprise extra
    row += 2
    ws[f'A{row}'] = 'PROFESSIONAL & ENTERPRISE EXTRA'
    apply_cell_style(ws[f'A{row}'], **create_header_style())
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    prof_extras = [
        '✅ Phone support',
        '✅ Priority bug fixing',
        '✅ Feature requests consideration',
        '✅ Monthly backups (backend)',
        '✅ Performance monitoring',
    ]
    
    for extra in prof_extras:
        ws[f'A{row}'] = extra
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
    
    # À la carte services
    row += 2
    ws[f'A{row}'] = 'À LA CARTE SERVICES'
    apply_cell_style(ws[f'A{row}'], **create_header_style())
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = 'Service'
    ws[f'B{row}'] = 'Prijs'
    ws[f'C{row}'] = 'Beschrijving'
    
    for col in [1, 2, 3]:
        cell = ws.cell(row=row, column=col)
        subheader_style = create_subheader_style()
        apply_cell_style(cell, **subheader_style)
    
    row += 1
    services = [
        ('Training sessie', 500, 'Max 10 personen, on-site of online'),
        ('Custom development', '90-150/uur', 'Maatwerk features'),
        ('Data migratie', '1,000-3,000', 'Van ander systeem'),
        ('On-premise deployment', 2500, 'Setup eigen server'),
        ('Extra integratie', '2,500-5,000', 'Per systeem'),
        ('Emergency support', '200/uur', 'Buiten support uren'),
        ('Code audit', 1500, 'Security & performance review'),
    ]
    
    for service, price, description in services:
        ws[f'A{row}'] = service
        ws[f'B{row}'] = price if isinstance(price, str) else price
        ws[f'C{row}'] = description
        
        if isinstance(price, (int, float)):
            ws[f'B{row}'].number_format = create_currency_format()
        
        ws.merge_cells(f'C{row}:E{row}')
        row += 1
    
    # Kolom breedtes
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15

def create_license_comparison(wb):
    """Sheet 8: License Models Comparison"""
    ws = wb.create_sheet("Licentie Modellen")
    
    # Header
    ws['A1'] = 'LICENTIE MODELLEN VERGELIJKING'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=COLORS['primary'])
    ws.merge_cells('A1:F1')
    
    # Model A: SaaS
    row = 3
    ws[f'A{row}'] = 'MODEL A: SaaS SUBSCRIPTION'
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    ws[f'A{row}'].font = Font(bold=True, color='FFFFFF', size=12)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    ws[f'A{row}'] = 'Tier'
    ws[f'B{row}'] = 'Prijs/gebruiker/maand'
    ws[f'C{row}'] = 'Min. Gebruikers'
    ws[f'D{row}'] = 'Target'
    
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        subheader_style = create_subheader_style()
        apply_cell_style(cell, **subheader_style)
    
    row += 1
    saas_tiers = [
        ('Starter', 8, 1, '1-5 gebruikers'),
        ('Professional', 15, 5, '6-25 gebruikers'),
        ('Enterprise', 25, 10, '25+ gebruikers'),
    ]
    
    for tier, price, min_users, target in saas_tiers:
        ws[f'A{row}'] = tier
        ws[f'B{row}'] = price
        ws[f'C{row}'] = min_users
        ws[f'D{row}'] = target
        
        ws[f'B{row}'].number_format = '€#,##0.00'
        row += 1
    
    row += 1
    ws[f'A{row}'] = 'Voordelen:'
    ws[f'A{row}'].font = Font(bold=True, color=COLORS['success'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws[f'A{row}'] = '+ Lage entry cost, + Schaalbaar, + Predictable kosten, + Includes updates'
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Nadelen:'
    ws[f'A{row}'].font = Font(bold=True, color=COLORS['danger'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws[f'A{row}'] = '- Ongoing kosten, - Totale TCO hoger op lange termijn'
    ws.merge_cells(f'A{row}:F{row}')
    
    # Model B: One-time License
    row += 3
    ws[f'A{row}'] = 'MODEL B: ONE-TIME LICENSE (PERPETUAL)'
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid')
    ws[f'A{row}'].font = Font(bold=True, color='FFFFFF', size=12)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    ws[f'A{row}'] = 'Package'
    ws[f'B{row}'] = 'Gebruikers'
    ws[f'C{row}'] = 'Eenmalige Fee'
    ws[f'D{row}'] = 'Jaarlijks Onderhoud'
    
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        subheader_style = create_subheader_style()
        apply_cell_style(cell, **subheader_style)
    
    row += 1
    license_packages = [
        ('Basis', '1-5', 2500, 500),
        ('Standaard', '6-15', 5000, 1000),
        ('Professional', '16-50', 10000, 2000),
        ('Enterprise', '50+', 15000, 3000),
    ]
    
    for package, users, fee, maintenance in license_packages:
        ws[f'A{row}'] = package
        ws[f'B{row}'] = users
        ws[f'C{row}'] = fee
        ws[f'D{row}'] = maintenance
        
        ws[f'C{row}'].number_format = create_currency_format()
        ws[f'D{row}'].number_format = create_currency_format()
        row += 1
    
    row += 1
    ws[f'A{row}'] = 'Voordelen:'
    ws[f'A{row}'].font = Font(bold=True, color=COLORS['success'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws[f'A{row}'] = '+ Lagere TCO op lange termijn, + Eenmalige investering, + Predictable kosten'
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Nadelen:'
    ws[f'A{row}'].font = Font(bold=True, color=COLORS['danger'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws[f'A{row}'] = '- Hogere upfront cost, - Onderhoud niet verplicht (maar aanbevolen)'
    ws.merge_cells(f'A{row}:F{row}')
    
    # Model C: Hybrid (AANBEVOLEN)
    row += 3
    ws[f'A{row}'] = 'MODEL C: HYBRID (AANBEVOLEN ⭐)'
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    ws[f'A{row}'].font = Font(bold=True, color='FFFFFF', size=12)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    ws[f'A{row}'] = 'Basis Setup Fee: €4,500 (eenmalig)'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Inclusief: Installatie, configuratie, training (4u), logo setup, data migratie'
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    ws[f'A{row}'] = 'Plus: €6/gebruiker/maand (min. 5 gebruikers)'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Inclusief: Hosting, updates, email support, backups, feature updates'
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 2
    ws[f'A{row}'] = 'Voordelen:'
    ws[f'A{row}'].font = Font(bold=True, color=COLORS['success'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws[f'A{row}'] = '+ Beste van beide werelden, + Lagere entry cost dan license, + All-inclusive support'
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    ws[f'A{row}'] = '+ Schaalbaar, + Predictable, + Professionele onboarding'
    ws.merge_cells(f'A{row}:F{row}')
    
    # TCO Comparison Chart
    row += 4
    ws[f'A{row}'] = '3-JAAR TCO VERGELIJKING (10 gebruikers)'
    apply_cell_style(ws[f'A{row}'], **create_header_style())
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    ws[f'A{row}'] = 'Model'
    ws[f'B{row}'] = 'Year 0'
    ws[f'C{row}'] = 'Year 1'
    ws[f'D{row}'] = 'Year 2'
    ws[f'E{row}'] = 'Year 3'
    ws[f'F{row}'] = 'Totaal'
    
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        subheader_style = create_subheader_style()
        apply_cell_style(cell, **subheader_style)
    
    row += 1
    # SaaS: 10 * €15 * 12 = €1,800/jaar
    ws[f'A{row}'] = 'SaaS (Professional)'
    ws[f'B{row}'] = 0
    ws[f'C{row}'] = 1800
    ws[f'D{row}'] = 1800
    ws[f'E{row}'] = 1800
    ws[f'F{row}'] = '=B{0}+C{0}+D{0}+E{0}'.format(row)
    
    row += 1
    # One-time: €5,000 + 3*€1,000 = €8,000
    ws[f'A{row}'] = 'One-time License'
    ws[f'B{row}'] = 5000
    ws[f'C{row}'] = 1000
    ws[f'D{row}'] = 1000
    ws[f'E{row}'] = 1000
    ws[f'F{row}'] = '=B{0}+C{0}+D{0}+E{0}'.format(row)
    
    row += 1
    # Hybrid: €4,500 + 3*(10*€6*12) = €4,500 + €2,160 = €6,660
    ws[f'A{row}'] = 'Hybrid ⭐'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = 4500
    ws[f'C{row}'] = 720
    ws[f'D{row}'] = 720
    ws[f'E{row}'] = 720
    ws[f'F{row}'] = '=B{0}+C{0}+D{0}+E{0}'.format(row)
    
    # Format all costs
    for r in range(row-2, row+1):
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{r}'].number_format = create_currency_format()
    
    # Highlight winner (Hybrid for 3 years)
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
        ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
    
    # Kolom breedtes
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18

# Main execution
if __name__ == '__main__':
    print("Genereren CH Construct Timesheet Excel Offerte...")
    
    wb = create_workbook()
    
    filename = 'CH-Construct-Timesheet-Commerciele-Offerte.xlsx'
    wb.save(filename)
    
    print(f"Excel bestand gegenereerd: {filename}")
    print(f"Aantal sheets: {len(wb.sheetnames)}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")
    print("\nOpen het bestand in Excel om alle interactieve calculators te gebruiken!")
